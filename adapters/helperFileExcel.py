# adapters/helperFileExcel.py
# JSON-BASED MEMORY (NO MULTIPLE FILES)

import os
import shutil
import json
import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from robot.api import logger
from datetime import datetime
import re
import time
import win32com.client





JSON_FILENAME = "internal_memory.json"


def _is_empty(v):
    return v is None or str(v).strip() == ""






def build_internal_memory(directory_files, root_dir):
    """
    Reads each Excel file
    Creates one JSON per Excel file
    Saves inside ROOT_DIR/Data/
    """

    processado_dir = os.path.join(directory_files, "Processado")

    # Create ROOT/Data folder
    data_dir = os.path.join(root_dir, "Data")
    os.makedirs(data_dir, exist_ok=True)

    print(f"Scanning folder: {processado_dir}")

    for fname in os.listdir(processado_dir):

        if fname.startswith("~$") or not fname.lower().endswith(".xlsm"):
            continue

        file_path = os.path.join(processado_dir, fname)
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.worksheets[0]

        file_memory = []
        file_has_match = False

        for excel_row in range(15, sheet.max_row + 1):

            global_id = sheet.cell(excel_row, 1).value
            empresa = sheet.cell(excel_row, 5).value
            req = sheet.cell(excel_row, 12).value

            if global_id not in (None, "") and req in (None, ""):

                file_has_match = True

                file_memory.append({
                    "FILENAME": file_path,
                    "COUNTER_STEP": empresa,
                    "ExcelRowIndex": excel_row,
                    "RequisitionNumber": ""
                })

    
        if file_has_match:

            json_filename = os.path.splitext(fname)[0] + ".json"
            json_path = os.path.join(data_dir, json_filename)

            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(file_memory, f, indent=4)

            print(f"Created JSON: {json_path}")

    
        else:
            ts = datetime.now().strftime("%d%m%Y%H%M")

            no_data_dir = os.path.join(root_dir, "No_Data_Files")
            os.makedirs(no_data_dir, exist_ok=True)

            shutil.move(
                file_path,
                os.path.join(no_data_dir, f"{os.path.splitext(fname)[0]}_{ts}.xlsm")
            )

            print(f"Moved (no data): {fname}")

    print("Internal memory creation completed.")


def create_plant_error_and_stop(
    empresa: str,
    tipo_servico: str,
    local_project_path: str
):
    """
    Creates internal memory table entry when Plant is invalid
    and updates the final analytical Excel report.
    Terminates execution afterwards.
    """

    error_entry = {
        "Etapa do processo": "Criação da Requisição de Compras SAP",
        "Data do processamento": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "Filial": empresa,
        "Descrição": "Planta inválida",
        "Curso": tipo_servico
    }

    error_dir = os.path.join(local_project_path, "ErrorHandle")
    os.makedirs(error_dir, exist_ok=True)

    error_json_path = os.path.join(error_dir, "internal_error_memory.json")

    if os.path.exists(error_json_path):
        with open(error_json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = []

    data.append(error_entry)

    with open(error_json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

    report_path = os.path.join(
        local_project_path,
        "Report",
        "Relatório Final_Analítico.xlsx"
    )


    if os.path.exists(report_path):
        wb = load_workbook(report_path)
        sheet = wb.active

        sheet.append([
            error_entry["Etapa do processo"],
            error_entry["Data do processamento"],
            error_entry["Filial"],
            error_entry["Descrição"],
            error_entry["Curso"]
        ])

        wb.save(report_path)
    else:
        raise FileNotFoundError(
            f"Report file not found at: {report_path}"
        )

    raise Exception("Planta inválida - execução interrompida.")







def generate_sorted_json_common(root_dir):
    """
    - Reads all JSON files from ROOT/Data
    - Removes Index
    - Removes ExcelRowIndex
    - Removes duplicates
    - Sorts data
    - Saves all sorted JSONs inside Data/sortedJson/
    """

    data_dir = os.path.join(root_dir, "Data")
    sorted_dir = os.path.join(data_dir, "sortedJson")

    os.makedirs(sorted_dir, exist_ok=True)

    for fname in os.listdir(data_dir):

        if not fname.lower().endswith(".json"):
            continue

        input_json = os.path.join(data_dir, fname)

        with open(input_json, "r", encoding="utf-8") as f:
            data = json.load(f)

        # ---------- REMOVE UNWANTED COLUMNS ----------
        cleaned_rows = []
        for r in data:
            cleaned_rows.append({
                "FILENAME": r.get("FILENAME"),
                "COUNTER_STEP": r.get("COUNTER_STEP"),
                "RequisitionNumber": r.get("RequisitionNumber", "")
            })

        # ---------- REMOVE DUPLICATES ----------
        unique_map = {}
        for r in cleaned_rows:
            key = (
                r.get("FILENAME"),
                r.get("COUNTER_STEP")
            )
            if key not in unique_map:
                unique_map[key] = r

        unique_rows = list(unique_map.values())

        # ---------- SORT ----------
        unique_rows.sort(
            key=lambda x: (
                x.get("FILENAME", ""),
                x.get("COUNTER_STEP", "")
            )
        )

        # ---------- SAVE ----------
        base_name = os.path.splitext(fname)[0]
        output_json = os.path.join(sorted_dir, f"sorted_{base_name}.json")

        with open(output_json, "w", encoding="utf-8") as f:
            json.dump(unique_rows, f, indent=4)

        print(f"Created: {output_json}")

    print("All files processed successfully.")





def build_header_aux_table_filewise(directory_files):
    """
    - Reads all filename JSONs inside Data/Filenames/
    - Each JSON contains list of Excel file paths
    - Reads C1:D12 from FIRST sheet only (index 0)
    - Detects Empresa from column E (starting row 15)
    - Saves JSON inside VHP or VPP folder accordingly
    """

    filenames_dir = os.path.join(directory_files, "Data", "Filenames")
    output_dir = os.path.join(directory_files, "Data", "Header_Aux_Table")

    os.makedirs(output_dir, exist_ok=True)

    if not os.path.isdir(filenames_dir):
        print("Filenames directory not found")
        return []

    json_files = [
        f for f in os.listdir(filenames_dir)
        if f.lower().endswith(".json")
    ]

    if not json_files:
        print("No filename JSON files found")
        return []

    for json_file in json_files:

        json_path = os.path.join(filenames_dir, json_file)

        with open(json_path, "r", encoding="utf-8") as f:
            file_paths = json.load(f)

        for file_path in file_paths:

            file_path = os.path.abspath(file_path)
            print(f"\n Processing: {file_path}")

            if not os.path.exists(file_path):
                print(f" File not found: {file_path}")
                continue

            wb = load_workbook(file_path, data_only=True)

            if not wb.worksheets:
                print("Workbook has no sheets")
                continue

            sheet = wb.worksheets[0]
            print(f"    Using Sheet: {sheet.title}")

            rows = []

            # -------------------------
            # Read C1:D12
            # -------------------------
            for r in range(1, 13):
                rows.append([
                    str(sheet.cell(row=r, column=3).value or ""),
                    str(sheet.cell(row=r, column=4).value or "")
                ])

            # -------------------------
            # Detect Empresa from column E (rows 15+)
            # -------------------------
            empresa_value = ""

            for r in range(15, sheet.max_row + 1):
                value = str(sheet.cell(row=r, column=5).value or "").strip().upper()

                if value in ["VHP", "VPP"]:
                    empresa_value = value
                    break

            if empresa_value == "VHP":
                company_dir = os.path.join(output_dir, "VHP")

            elif empresa_value == "VPP":
                company_dir = os.path.join(output_dir, "VPP")

            else:
                print(f"⚠ Empresa not detected in {file_path}")
                company_dir = output_dir

            os.makedirs(company_dir, exist_ok=True)

            file_name = os.path.splitext(os.path.basename(file_path))[0]

            output_json = os.path.join(
                company_dir,
                f"{file_name}_Header_Aux_Table.json"
            )

            with open(output_json, "w", encoding="utf-8") as f:
                json.dump({
                    "file": file_path,
                    "rows": rows
                }, f, indent=4, ensure_ascii=False)

            print(f"✅ Created: {output_json}")

    print("\n🎯 Header Aux Extraction Completed")
    return True





def extract_all_header_fields(directory_files, company_type):
    """
    - Reads JSON files inside Data/Header_Aux_Table/VHP or VPP
    - Extracts header key-value pairs
    - Returns list of dictionaries file-wise
    """

    # Select correct company folder
    header_dir = os.path.join(
        directory_files,
        "Data",
        "Header_Aux_Table",
        company_type
    )

    if not os.path.isdir(header_dir):
        logger.console(f" Header directory not found: {header_dir}")
        return []

    extracted_list = []

    json_files = [
        f for f in os.listdir(header_dir)
        if f.lower().endswith(".json")
    ]

    if not json_files:
        logger.console(" No Header_Aux_Table JSON files found")
        return []

    for json_file in json_files:

        json_path = os.path.join(header_dir, json_file)

        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        rows = data.get("rows", [])
        file_path = data.get("file", "")
        file_name = os.path.basename(file_path)

        file_dict = {}

        for row in rows:

            if len(row) < 2:
                continue

            key = str(row[0]).strip()
            value = str(row[1]).strip() if row[1] else ""

            if key:
                file_dict[key] = value

        extracted_list.append({
            "file": file_name,
            "headers": file_dict
        })

    logger.console(f" Header fields extracted for {company_type}")

    return extracted_list






def extract_filenames_per_sorted_json(root_dir):
    """
    - Reads all sorted_*.json inside Data/sortedJson
    - Extracts FILENAME values
    - Creates file-wise filename JSONs inside Data/Filenames/
    """

    sorted_dir = os.path.join(root_dir, "Data", "sortedJson")
    output_dir = os.path.join(root_dir, "Data", "Filenames")

    os.makedirs(output_dir, exist_ok=True)

    if not os.path.exists(sorted_dir):
        print("sortedJson directory not found")
        return

    for fname in os.listdir(sorted_dir):

        if not fname.lower().endswith(".json"):
            continue

        input_json = os.path.join(sorted_dir, fname)

        with open(input_json, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Extract filenames
        filenames = [
            row.get("FILENAME")
            for row in data
            if row.get("FILENAME")
        ]

        base_name = os.path.splitext(fname)[0]
        output_json = os.path.join(
            output_dir,
            f"filenames_{base_name}.json"
        )

        with open(output_json, "w", encoding="utf-8") as f:
            json.dump(filenames, f, indent=4)

        # print(f"Created: {output_json}")
        print(f"Total filenames: {len(filenames)}")

    print("Filename extraction completed.")




def extract_all_header_fields_from_finaltable(directory_files, company_type):
    """
    - Reads JSON files from:
        Data/FinalTable/VPPfinaltable   OR
        Data/FinalTable/VHPfinaltable
    - company_type must be 'VPP' or 'VHP'
    - Returns all final table entries file-wise
    """

    company_type = company_type.strip().upper()

    finaltable_root = os.path.join(directory_files, "Data", "FinalTable")

    if company_type == "VPP":
        target_dir = os.path.join(finaltable_root, "VPPfinaltable")
    elif company_type == "VHP":
        target_dir = os.path.join(finaltable_root, "VHPfinaltable")
    else:
        logger.console(f" Invalid company type: {company_type}")
        return []

    if not os.path.isdir(target_dir):
        logger.console(f" Directory not found: {target_dir}")
        return []

    extracted_list = []

    for fname in os.listdir(target_dir):
        if not fname.lower().endswith(".json"):
            continue

        json_path = os.path.join(target_dir, fname)

        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Each file contains list → extend full content
        extracted_list.extend(data)

    logger.console(f" Final table extracted successfully from {company_type}")

    return extracted_list


# READ A14:L1048576 (EMPLOYEES) FROM VISIBLE SHEET → SAVE JSON



def _json_safe(value):
    """Convert datetime to string for JSON"""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return value




def extract_employees_filewise(directory_files):
    """
    - Reads all filename JSONs inside Data/Filenames/
    - Each JSON contains list of Excel file paths
    - Uses Row 14 as header
    - Reads A15:L(last row) as employee data
    - Uses sheet index 0 only
    - Saves separate JSON per file
    - Output folder:
        directory_files/Data/Employees/
        directory_files/Data/Employees/VHP/
        directory_files/Data/Employees/VPP/
    """

    # -----------------------------
    # Define Paths
    # -----------------------------
    filenames_dir = os.path.join(directory_files, "Data", "Filenames")
    output_dir = os.path.join(directory_files, "Data", "Employees")

    os.makedirs(output_dir, exist_ok=True)

    if not os.path.isdir(filenames_dir):
        logger.console(" Filenames directory not found")
        return []

    json_files = [
        f for f in os.listdir(filenames_dir)
        if f.lower().endswith(".json")
    ]

    if not json_files:
        logger.console(" No filename JSON files found")
        return []

    # -----------------------------
    # Process Each Filename JSON
    # -----------------------------
    for json_file in json_files:

        json_path = os.path.join(filenames_dir, json_file)

        with open(json_path, "r", encoding="utf-8") as f:
            file_paths = json.load(f)

        for file_path in file_paths:

            file_path = os.path.abspath(file_path)
            logger.console(f"\n Processing: {file_path}")

            if not os.path.exists(file_path):
                logger.console(f" FILE NOT FOUND: {file_path}")
                continue

            try:
                wb = load_workbook(file_path, data_only=True)
            except Exception as e:
                logger.console(f" Error opening file: {e}")
                continue

            # -----------------------------
            # Select Sheet Index 0 Only
            # -----------------------------
            if not wb.worksheets:
                logger.console(" Workbook has no sheets")
                continue

            sheet = wb.worksheets[0]

            max_col = 12  # A to L

            # -------------------------
            # Read Header Row (Row 14)
            # -------------------------
            headers = []

            for c in range(1, max_col + 1):
                header_value = sheet.cell(row=14, column=c).value
                headers.append(str(header_value or "").strip())

            employees_rows = []

            # -------------------------
            # Read Data Rows (Row 15+)
            # -------------------------
            for r in range(15, sheet.max_row + 1):

                row_dict = {}
                is_empty = True

                for c in range(1, max_col + 1):
                    cell_value = sheet.cell(row=r, column=c).value
                    value = str(cell_value or "").strip()

                    if value != "":
                        is_empty = False

                    key = headers[c - 1] if headers[c - 1] != "" else f"Column_{c}"
                    row_dict[key] = value

                if not is_empty:
                    employees_rows.append(row_dict)

            logger.console(f" EMPLOYEE ROWS FOUND: {len(employees_rows)}")

            # -------------------------
            # Detect Empresa Type
            # -------------------------
            empresa_set = {
                str(emp.get("Empresa", "")).strip().upper()
                for emp in employees_rows
            }

            company_folder = ""

            if "VHP" in empresa_set:
                company_folder = "VHP"
            elif "VPP" in empresa_set:
                company_folder = "VPP"

            # -------------------------
            # Decide Output Directory
            # -------------------------
            if company_folder:
                output_company_dir = os.path.join(output_dir, company_folder)
            else:
                output_company_dir = output_dir

            os.makedirs(output_company_dir, exist_ok=True)

            # -------------------------
            # Save File-wise JSON
            # -------------------------
            file_name = os.path.splitext(os.path.basename(file_path))[0]

            output_json = os.path.join(
                output_company_dir,
                f"{file_name}_Employees.json"
            )

            with open(output_json, "w", encoding="utf-8") as f:
                json.dump({
                    "file": file_path,
                    "employees": employees_rows
                }, f, indent=4, ensure_ascii=False)

            logger.console(f"✅ Created: {output_json}")

    logger.console("\n Employees Extraction Completed Successfully")
    return True


EMPRESA_COLUMN_INDEX = 4  # Column E

def build_final_table_filewise(directory_files):

    data_root = os.path.join(directory_files, "Data")
    sorted_dir = os.path.join(data_root, "sortedJson")

    employees_root_dir = os.path.join(data_root, "Employees")
    final_root_dir = os.path.join(data_root, "FinalTable")

    os.makedirs(final_root_dir, exist_ok=True)

    if not os.path.exists(sorted_dir):
        logger.console(" sortedJson directory not found")
        return False

    if not os.path.exists(employees_root_dir):
        logger.console(" Employees directory not found")
        return False

    for sorted_file in os.listdir(sorted_dir):

        if not sorted_file.endswith(".json"):
            continue

        sorted_json_path = os.path.join(sorted_dir, sorted_file)

        base_name = sorted_file.replace("sorted_", "")

        logger.console(f"🔍 Processing: {base_name}")

        with open(sorted_json_path, "r", encoding="utf-8") as f:
            sorted_data = json.load(f)

        # Prepare final tables
        final_tables = {
            "VPP": [],
            "VHP": []
        }

        for item in sorted_data:

            counter_step = str(item.get("COUNTER_STEP", "")).strip().upper()
            filename = item.get("FILENAME")

            if counter_step not in ["VPP", "VHP"]:
                continue

            # -----------------------------
            # Select correct Employees folder
            # -----------------------------
            employees_json_path = os.path.join(
                employees_root_dir,
                counter_step,
                base_name.replace(".json", "_Employees.json")
            )

            if not os.path.exists(employees_json_path):
                logger.console(f" Employees JSON not found for {base_name} in {counter_step}")
                continue

            with open(employees_json_path, "r", encoding="utf-8") as f:
                employees_data = json.load(f)

            employees_list = employees_data.get("employees", [])

            matched_employees = []

            for emp in employees_list:

                empresa_value = str(emp.get("Empresa", "")).strip().upper()

                if empresa_value == counter_step:
                    matched_employees.append(emp)

            if matched_employees:
                final_tables[counter_step].append({
                    "file": filename,
                    "employeeCount": len(matched_employees),
                    "counter_step": counter_step,
                    "employees": matched_employees
                })

        # ----------------------------
        # Save VPP and VHP separately
        # ----------------------------
        for step in ["VPP", "VHP"]:

            if not final_tables[step]:
                continue

            step_dir = os.path.join(final_root_dir, f"{step}finaltable")
            os.makedirs(step_dir, exist_ok=True)

            output_filename = base_name.replace(
                ".json", f"_{step}_FinalTable.json"
            )

            output_path = os.path.join(step_dir, output_filename)

            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(final_tables[step], f, indent=4, ensure_ascii=False)

            logger.console(f" FinalTable created: {output_path}")

    return True

# adapters/company_costs_helper.py
# READ A1:B6 FROM ALL VISIBLE SHEETS OF ALL FILES → STRUCTURED JSON



def extract_company_costs_filewise(directory_files):
    """
    - Reads all filename JSONs inside Data/Filenames/
    - Each JSON contains list of Excel file paths
    - For each Excel file:
        - Reads A1:B6 from FIRST sheet only (index 0)
        - Saves separate JSON per file
    - Output folder:
        directory_files/Data/Company_Costs/
    """

    # -----------------------------
    # Define Paths
    # -----------------------------

    filenames_dir = os.path.join(directory_files, "Data", "Filenames")
    data_dir = os.path.join(directory_files, "Data", "Company_Costs")

    os.makedirs(data_dir, exist_ok=True)

    print(f"\n Output Directory: {data_dir}")

    # -----------------------------
    # Validate Filenames Directory
    # -----------------------------
    if not os.path.isdir(filenames_dir):
        print(" Filenames directory not found")
        return []

    json_files = [
        f for f in os.listdir(filenames_dir)
        if f.lower().endswith(".json")
    ]

    if not json_files:
        print(" No filename JSON files found")
        return []

    print(f" Total Filename JSONs Found: {len(json_files)}")

    # -----------------------------
    # Process Each Filename JSON
    # -----------------------------
    for json_file in json_files:

        json_path = os.path.join(filenames_dir, json_file)

        with open(json_path, "r", encoding="utf-8") as f:
            file_paths = json.load(f)

        for file_path in file_paths:

            file_path = os.path.abspath(file_path)
            print(f"\n Processing: {file_path}")

            if not os.path.exists(file_path):
                print(f" File not found: {file_path}")
                continue

            try:
                wb = load_workbook(file_path, data_only=True)
            except Exception as e:
                print(f" Error opening file: {e}")
                continue

            # -----------------------------
            # SELECT FIRST SHEET (INDEX 0)
            # -----------------------------
            if not wb.worksheets:
                print(" Workbook has no sheets.")
                continue

            sheet = wb.worksheets[0]
            print(f"    Using Sheet: {sheet.title}")

            sheet_data = []

            # Read A1:B6
            for r in range(1, 7):
                sheet_data.append([
                    sheet.cell(row=r, column=1).value,
                    sheet.cell(row=r, column=2).value
                ])

            file_result = [{
                "sheet": sheet.title,
                "data": sheet_data
            }]

            # -----------------------------
            # Create Output JSON (file-wise)
            # -----------------------------
            file_name = os.path.splitext(os.path.basename(file_path))[0]

            output_json = os.path.join(
                data_dir,
                f"{file_name}_Company_Costs.json"
            )

            with open(output_json, "w", encoding="utf-8") as f:
                json.dump(file_result, f, indent=4)

            print(f" Created: {output_json}")

    print("\n🎯 Company Costs Extraction Completed Successfully")
    return True

# adapters/currency_helper.py
# READ E4 (Currency) FROM VISIBLE SHEET → SAVE JSON



def extract_currency_filewise(directory_files):
    """
    - Reads all filename JSONs inside Data/Filenames/
    - Each JSON contains list of Excel file paths
    - Reads cell E4 from first sheet (index 0)
    - Saves separate JSON per file
    - Output folder:
        directory_files/Data/Currency/
    """

    # -----------------------------
    # Define Paths
    # -----------------------------
    filenames_dir = os.path.join(directory_files, "Data", "Filenames")
    output_dir = os.path.join(directory_files, "Data", "Currency")

    os.makedirs(output_dir, exist_ok=True)

    # -----------------------------
    # Validate Filenames Directory
    # -----------------------------
    if not os.path.isdir(filenames_dir):
        logger.console(" Filenames directory not found")
        return []

    json_files = [
        f for f in os.listdir(filenames_dir)
        if f.lower().endswith(".json")
    ]

    if not json_files:
        logger.console(" No filename JSON files found")
        return []

    # -----------------------------
    # Process Each Filename JSON
    # -----------------------------
    for json_file in json_files:

        json_path = os.path.join(filenames_dir, json_file)

        with open(json_path, "r", encoding="utf-8") as f:
            file_paths = json.load(f)

        for file_path in file_paths:

            file_path = os.path.abspath(file_path)
            logger.console(f"\n🔄 Processing: {file_path}")

            if not os.path.exists(file_path):
                logger.console(f"⚠ FILE NOT FOUND: {file_path}")
                continue

            try:
                wb = load_workbook(file_path, data_only=True)
            except Exception as e:
                logger.console(f"❌ Error opening file: {e}")
                continue

            # -----------------------------
            # Select First Sheet (Index 0)
            # -----------------------------
            if not wb.worksheets:
                logger.console("⚠ Workbook has no sheets")
                continue

            sheet = wb.worksheets[0]

            # -----------------------------
            # Read Currency from E4
            # -----------------------------
            currency_value = sheet.cell(row=4, column=5).value  # E4

            # -----------------------------
            # Create Output JSON
            # -----------------------------
            file_name = os.path.splitext(os.path.basename(file_path))[0]

            output_json = os.path.join(
                output_dir,
                f"{file_name}_Currency.json"
            )

            with open(output_json, "w", encoding="utf-8") as f:
                json.dump({
                    "file": file_path,
                    "currency": str(currency_value or "")
                }, f, indent=4, ensure_ascii=False)

            logger.console(f"✅ Created: {output_json}")

    logger.console("\n🎯 Currency Extraction Completed")
    return True





TABLE_ID = (
    "wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100"
    "/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301"
    "/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT7"
    "/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101"
    "/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000"
    "/tblSAPLMEACCTVIDYN_1000TC"
)

def get_active_session():
    """Always fetches the active SAP session reliably."""
    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        for i in range(application.Children.Count):
            connection = application.Children(i)
            for j in range(connection.Children.Count):
                session = connection.Children(j)
                if not session.Busy:
                    return session
        raise Exception("No active SAP session found")
    except Exception as e:
        raise Exception(f"SAP session error: {str(e)}")

def scroll_sap_table(scroll_pos):
    try:
        session = get_active_session()
        table = session.findById(TABLE_ID)
        table.verticalScrollbar.position = int(scroll_pos)
        print(f"Scrolled to: {scroll_pos}")
    except Exception as e:
        print(f"SCROLL ERROR: {str(e)}")
        raise
def extract_hydro_plant(value: str, empresa: str) -> str:
    print("DEBUG VALUE RECEIVED:", value)
    print("DEBUG EMPRESA:", empresa)

    if not value:
        return ""

    parts = value.split("|")

    # VPP → first value
    if empresa.strip().upper() == "VPP":
        return parts[0].strip()

    # VHP → second block first value
    if empresa.strip().upper() == "VHP":
        if len(parts) > 1:
            second_part = parts[1]
            sub_parts = second_part.split(";")
            return sub_parts[0].strip()

    return ""




def wait_until_file_is_available(path, timeout=30):
    start = time.time()

    while True:
        try:
            with open(path, "a"):
                return
        except PermissionError:
            if time.time() - start > timeout:
                raise PermissionError(f"File locked: {path}")
            time.sleep(1)


def update_excel_and_report(filename, sap_status, empresa, tipo_servico, project_dir, report_dir):

    # Extract PR number
    match = re.search(r"\d+", sap_status)
    pr_number = match.group() if match else ""

    # -----------------------------------------
    # STEP 1 — Read Excel from Processado
    # -----------------------------------------
    source_path = os.path.join(project_dir, filename)

    if not os.path.exists(source_path):
        raise FileNotFoundError(f"Excel file not found: {source_path}")

    wait_until_file_is_available(source_path)

    wb = load_workbook(source_path, data_only=True)
    sheet = wb.worksheets[0]

    # -----------------------------------------
    # STEP 2 — Read A15:L into memory table
    # -----------------------------------------
    memory_table = []

    for row in sheet.iter_rows(min_row=15, min_col=1, max_col=12):
        row_values = [cell.value for cell in row]
        memory_table.append(row_values)

    wb.close()

    # -----------------------------------------
    # STEP 3 — Update column K in memory table
    # -----------------------------------------
    for row in memory_table:
        if len(row) >= 11:
            row[10] = pr_number   # Column K index = 10

    # -----------------------------------------
    # STEP 4 — Create report row
    # -----------------------------------------
    report_row = [
        "Requisição foi criada",
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        empresa,
        sap_status,
        tipo_servico
    ]

    # -----------------------------------------
    # STEP 5 — Open analytical report
    # -----------------------------------------
    report_path = os.path.join(
        report_dir,
        "Report",
        "Relatório Final_Analítico.xlsx"
    )

    if not os.path.exists(report_path):
        raise FileNotFoundError(f"Report file not found: {report_path}")

    wait_until_file_is_available(report_path)

    report_wb = load_workbook(report_path)
    report_sheet = report_wb.active

    # -----------------------------------------
    # STEP 6 — Append report row
    # -----------------------------------------
    report_sheet.append(report_row)

    report_wb.save(report_path)
    report_wb.close()

    return {
        "pr_number": pr_number,
        "memory_table_rows": len(memory_table)
    }





def handle_attachment_status(local_project, empresa, tipo_servico, sap_status):

    # Internal memory row (same for success or failure)
    etapa = "Anexar arquivo no SAP"
    data_proc = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    excel_path = os.path.join(
        local_project,
        "Report",
        "Relatório Final_Analítico.xlsx"
    )

    wb = load_workbook(excel_path)
    sheet = wb.active

    next_row = sheet.max_row + 1

    sheet.cell(next_row, 1).value = etapa
    sheet.cell(next_row, 2).value = data_proc
    sheet.cell(next_row, 3).value = empresa
    sheet.cell(next_row, 4).value = sap_status
    sheet.cell(next_row, 5).value = tipo_servico

    wb.save(excel_path)

    return True



def create_error_record(
    etapa_processo: str,
    descricao: str,
    empresa: str,
    tipo_servico: str,
    local_project_path: str,
    error_file_name: str
):
    """
    Generic error handler that:
    - Creates internal memory table
    - Saves into specific JSON file
    - Updates Excel report
    """

    error_entry = {
        "Etapa do processo": etapa_processo,
        "Data do processamento": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "Filial": empresa,
        "Descrição": descricao,
        "Curso": tipo_servico
    }

    # Create ErrorHandle folder
    error_dir = os.path.join(local_project_path, "ErrorHandle")
    os.makedirs(error_dir, exist_ok=True)

    # Create specific JSON file
    error_json_path = os.path.join(error_dir, error_file_name)

    if os.path.exists(error_json_path):
        with open(error_json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = []

    data.append(error_entry)

    with open(error_json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

    # Update Excel report
    report_path = os.path.join(
        local_project_path,
        "Report",
        "Relatório Final_Analítico.xlsx"
    )

    if os.path.exists(report_path):
        wb = load_workbook(report_path)
        sheet = wb.active

        sheet.append([
            error_entry["Etapa do processo"],
            error_entry["Data do processamento"],
            error_entry["Filial"],
            error_entry["Descrição"],
            error_entry["Curso"]
        ])

        wb.save(report_path)

    return error_entry



COLUMNS = ["Etapa do processo", "Data do processamento", "Filial", "Descrição", "Curso"]


def _get_timestamp():
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")

def _write_to_report(row_data, root_dir):
    """Write a row to the Excel report file dynamically using root_dir."""

    try:
        report_path = os.path.join(
            root_dir,
            "Report",
            "Relatório Final_Analítico.xlsx"
        )

        # Ensure directory exists
        os.makedirs(os.path.dirname(report_path), exist_ok=True)

        # -----------------------------------------
        # Open or create workbook
        # -----------------------------------------
        if os.path.exists(report_path):
            wb = openpyxl.load_workbook(report_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active

            # Add header row (only once when file is created)
            ws.append([
                "Etapa",
                "Data",
                "Filial",
                "Descricao",
                "Curso"
            ])

        # -----------------------------------------
        # Append data row (THIS is the important part)
        # -----------------------------------------
        ws.append([
            row_data.get("etapa", ""),
            row_data.get("data", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            row_data.get("filial", ""),
            row_data.get("descricao", ""),
            row_data.get("curso", "")
        ])

        # -----------------------------------------
        # Save file
        # -----------------------------------------
        wb.save(report_path)
        wb.close()

        print(f"[INFO] Report updated: {row_data.get('descricao', '')}")

    except Exception as e:
        print(f"[ERROR] Failed to write report: {str(e)}")
        raise

# ============================================================
# Step 5.3.2.2.10 — Invalid Plant Error
# ============================================================
def create_invalid_plant_error(empresa, tipo_servico, root_dir):
    """Called when SAP Plant is invalid."""
    row_data = {
        "etapa": "Criação da Requisição de Compras SAP",
        "data": _get_timestamp(),
        "filial": empresa,
        "descricao": "Planta inválida",
        "curso": tipo_servico
    }
    _write_to_report(row_data, root_dir)
    print(f"[ERROR HANDLER] Invalid Plant error logged for: {empresa}")


# ============================================================
# Step 5.3.2.2.11 — Invalid Storage Error
# ============================================================
def create_invalid_storage_error(empresa, tipo_servico, root_dir):
    """Called when SAP Storage Location is invalid."""
    row_data = {
        "etapa": "Criação da Requisição de Compras SAP",
        "data": _get_timestamp(),
        "filial": empresa,
        "descricao": "Código de armazenamento inválido",
        "curso": tipo_servico
    }
    _write_to_report(row_data, root_dir)
    print(f"[ERROR HANDLER] Invalid Storage error logged for: {empresa}")


# ============================================================
# Step 5.3.2.2.15 — CNPJ Not Found Error
# ============================================================
def create_cnpj_error(empresa, tipo_servico, sap_message, root_dir):
    """Called when CNPJ is not found in SAP vendor search."""
    row_data = {
        "etapa": "Criando requisição de compras",
        "data": _get_timestamp(),
        "filial": empresa,
        "descricao": f"Erro de CNPJ: {sap_message}",
        "curso": tipo_servico
    }
    _write_to_report(row_data, root_dir)
    print(f"[ERROR HANDLER] CNPJ error logged for: {empresa} — {sap_message}")


# ============================================================
# Step 5.3.2.2.16 — Vendor Blocked / Quantity / Service Approver Error
# ============================================================
def create_vendor_error(empresa, tipo_servico, sap_message, root_dir):
    """Called when vendor is blocked, quantity missing, or service approver error."""
    row_data = {
        "etapa": "Desired Vendor screen",
        "data": _get_timestamp(),
        "filial": empresa,
        "descricao": f"Erro do CNPJ: {sap_message}",
        "curso": tipo_servico
    }
    _write_to_report(row_data, root_dir)
    print(f"[ERROR HANDLER] Vendor error logged for: {empresa} — {sap_message}")


# ============================================================
# Step 5.3.2.14 — Account Assignment Errors
# ============================================================
def create_account_assignment_error(empresa, tipo_servico, sap_message, root_dir):
    """Called when account assignment errors occur after filling employee table."""

    if "Purchasing across company codes" in sap_message:
        etapa = "Creating Purchase Requisition"
        descricao = f"Erro na coluna 'CDC': {sap_message}"
    elif "Cost center" in sap_message:
        etapa = "Creating Purchase Requisition"
        descricao = f"Erro na coluna 'CDC': {sap_message}"
    elif "account" in sap_message.lower():
        etapa = "Creating Purchase Requisition"
        descricao = f"Erro na Conta Razão: {sap_message}"
    elif "Sum of percentages" in sap_message:
        etapa = "Creating Purchase Requisition"
        descricao = f"Erro na coluna 'Porcentagem': {sap_message}"
    else:
        etapa = "Creating Purchase Requisition"
        descricao = f"Erro desconhecido: {sap_message}"

    row_data = {
        "etapa": etapa,
        "data": _get_timestamp(),
        "filial": empresa,
        "descricao": descricao,
        "curso": tipo_servico
    }
    _write_to_report(row_data, root_dir)
    print(f"[ERROR HANDLER] Account assignment error logged: {descricao}")


# ============================================================
# Step 5.3.2.31 / 5.3.2.32 — Attachment Success or Failure
# ============================================================
def create_attachment_record(empresa, tipo_servico, sap_message, root_dir):
    """Called after SAP attachment attempt — success or failure."""
    row_data = {
        "etapa": "Anexar arquivo no SAP",
        "data": _get_timestamp(),
        "filial": empresa,
        "descricao": sap_message,
        "curso": tipo_servico
    }
    _write_to_report(row_data, root_dir)
    print(f"[ERROR HANDLER] Attachment record logged: {sap_message}")


# ============================================================
# Step 5.3.2.33 — SAP Cannot Open
# ============================================================
def create_sap_open_error(empresa, tipo_servico, root_dir):
    """Called when SAP cannot be opened."""
    row_data = {
        "etapa": "Abrir SAP",
        "data": _get_timestamp(),
        "filial": empresa,
        "descricao": "Não foi possível abrir o SAP",
        "curso": tipo_servico
    }
    _write_to_report(row_data, root_dir)
    print(f"[ERROR HANDLER] SAP open error logged for: {empresa}")


# ============================================================
# Step 5.3.3 — No Employees Found
# ============================================================
def create_no_employees_error(empresa, tipo_servico, root_dir):
    """Called when no employees are found for a company."""
    row_data = {
        "etapa": "Verificando Dados",
        "data": _get_timestamp(),
        "filial": empresa,
        "descricao": f"{empresa}: Não há dados.",
        "curso": tipo_servico
    }
    _write_to_report(row_data, root_dir)
    print(f"[ERROR HANDLER] No employees error logged for: {empresa}")


# ============================================================
# Step 5.3.2.19 / 5.3.2.23 — PR Created Successfully
# ============================================================
def create_pr_success_record(empresa, tipo_servico, sap_message, root_dir):
    """Called when Purchase Requisition is created successfully."""
    row_data = {
        "etapa": "Requisição foi criada",
        "data": _get_timestamp(),
        "filial": empresa,
        "descricao": sap_message,
        "curso": tipo_servico
    }
    _write_to_report(row_data, root_dir)
    print(f"[ERROR HANDLER] PR success record logged: {sap_message}")